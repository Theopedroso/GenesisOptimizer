"""
Gerador de Relatório Excel Profissional — v2.0
Genesis Optimizer · Exportação completa em memória.

Abas:
  1. Resumo Executivo    — KPIs biológicos + EPEF + econômico
  2. Programa Nutricional — composição por fase + perfil nutricional
  3. Análise Econômica   — break-even feed+total · IOFC · IOAC
  4. Aminoácidos & EB    — razões AA todas as fases + balanço eletrolítico
  5. Frota               — escalas de produção
  6. Curva Crescimento   — tabela diária/semanal estimada
  7. Shadow Prices       — restrições ativas LP
  8. Parâmetros          — preços, programa, custos extras
"""

from io import BytesIO
import numpy as np


def gerar_relatorio_excel(resultado: dict, programa: dict, precos: dict,
                          custos_extras: dict | None = None,
                          n_aves: int = 100_000,
                          mortalidade_pct: float = 3.0,
                          idade_d: int = 45) -> bytes:
    """
    Gera relatório Excel completo em memória.
    Retorna bytes prontos para download via Streamlit.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b""

    if custos_extras is None:
        custos_extras = {"pinto": 0.45, "medicacao": 0.05, "energia": 0.04,
                         "mao_obra": 0.07, "depreciacao": 0.10, "outros": 0.04}

    # ── Paleta ──────────────────────────────────────────────────────────────
    C_HEADER  = "0D2137"
    C_TITULO  = "0E3A5A"
    C_ACENTO  = "E07B00"
    C_VERDE   = "0A5C2A"
    C_AZUL2   = "1B6CA8"
    C_CLARO   = "D6EAF8"
    C_VERDE2  = "D5F5E3"
    C_AMBER   = "FEF9E7"
    C_CINZA   = "F2F3F4"
    C_BRANCO  = "FFFFFF"
    C_RED     = "FADBD8"
    C_ORANGE  = "FDEBD0"

    def _font(bold=True, cor="FFFFFF", sz=11, mono=False):
        name = "Courier New" if mono else "Calibri"
        return Font(name=name, bold=bold, color=cor, size=sz)

    def _fill(hex_cor):
        return PatternFill("solid", fgColor=hex_cor)

    def _border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def _align(h="center"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    def _titulo_aba(ws, titulo, subtitulo="", ncols=12):
        rng = f"A1:{get_column_letter(ncols)}1"
        ws.merge_cells(rng)
        c = ws["A1"]
        c.value = titulo
        c.font = Font("Calibri", bold=True, size=15, color="FFFFFF")
        c.fill = _fill(C_HEADER)
        c.alignment = _align()
        ws.row_dimensions[1].height = 28
        if subtitulo:
            rng2 = f"A2:{get_column_letter(ncols)}2"
            ws.merge_cells(rng2)
            c2 = ws["A2"]
            c2.value = subtitulo
            c2.font = Font("Calibri", size=10, color="555555")
            c2.alignment = _align()
            ws.row_dimensions[2].height = 16

    def _hdr(ws, row, cols, bg=C_TITULO, txt="FFFFFF", sz=10):
        for ci, t in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=t)
            c.font = Font("Calibri", bold=True, color=txt, size=sz)
            c.fill = _fill(bg)
            c.alignment = _align()
            c.border = _border()

    def _sec(ws, row, label, ncols=10, bg=C_AZUL2):
        rng = f"A{row}:{get_column_letter(ncols)}{row}"
        ws.merge_cells(rng)
        c = ws[f"A{row}"]
        c.value = label
        c.font = Font("Calibri", bold=True, color="FFFFFF", size=10)
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        return row + 1

    def _val(ws, row, col, v, fmt=None, bg=None, bold=False, align="center"):
        c = ws.cell(row=row, column=col, value=v)
        c.border = _border()
        c.alignment = _align(align)
        if fmt:
            c.number_format = fmt
        if bg:
            c.fill = _fill(bg)
        if bold:
            c.font = Font("Calibri", bold=True, size=10)
        return c

    def _bg_for(v, good_min, good_max=None, invert=False):
        if v is None:
            return C_CINZA
        try:
            fv = float(v)
        except Exception:
            return C_BRANCO
        if good_max is None:
            ok = (fv >= good_min) if not invert else (fv <= good_min)
        else:
            ok = good_min <= fv <= good_max
        return C_VERDE2 if ok else C_RED

    FASES  = ["starter", "grower", "finisher1", "finisher2"]
    F_NOME = ["Starter (0-12d)", "Grower (12-24d)", "Finisher 1 (24-36d)", "Finisher 2 (36-45d)"]
    F_SHT  = ["Starter", "Grower", "Fin.1", "Fin.2"]

    form_res = resultado.get("formulacao", {})

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════
    # ABA 1 · RESUMO EXECUTIVO
    # ════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumo Executivo"
    ws1.sheet_view.showGridLines = False
    _titulo_aba(ws1, "GENESIS OPTIMIZER  ·  Resumo Executivo v4.0",
                f"RSM Genesis42d_G · LP HiGHS · Programa: ME{int(programa.get('grower',{}).get('ame_n',3050))} Lys{programa.get('grower',{}).get('dig_lys',1.20):.3f}",
                ncols=8)

    from modules.optimizer import calcular_epef, custo_producao_total, escalas_frota
    desemp = resultado.get("desempenho", {})
    epef_d = calcular_epef(resultado, int(idade_d), float(mortalidade_pct))
    cprod  = custo_producao_total(resultado, custos_extras)
    frota  = escalas_frota(resultado, int(n_aves), float(mortalidade_pct))
    custo_ext = sum(custos_extras.values())

    row = 4
    row = _sec(ws1, row, "DESEMPENHO BIOLÓGICO  ·  RSM Genesis42d_G", ncols=8)
    _hdr(ws1, row, ["Indicador", "Valor", "Unidade", "Meta/Referência", "Status"], sz=9)
    dados_bio = [
        ("Peso Vivo RSM (ref ~49d)",    desemp.get("peso_vivo_kg"),     "kg",    ">3.5",  desemp.get("peso_vivo_kg", 0) >= 3.5),
        (f"Peso Vivo Estimado {idade_d}d", epef_d["peso_estimado_kg"],  "kg",    ">3.5",  epef_d["peso_estimado_kg"] >= 3.5),
        ("Conv. Alimentar (CAA)",        desemp.get("caa"),             "kg/kg", "<1.90", desemp.get("caa", 9) <= 1.90),
        ("CAA Ajustado (idade abate)",   epef_d["caa_ajustado"],        "kg/kg", "<1.95", epef_d["caa_ajustado"] <= 1.95),
        ("Ganho Diário",                 epef_d["ganho_diario_g"],      "g/d",   ">85",   epef_d["ganho_diario_g"] >= 85),
        ("Consumo Diário",               epef_d["consumo_diario_g"],    "g/d",   "—",     True),
        ("Carcaça WOG",                  desemp.get("carcaca_kg"),      "kg",    "—",     True),
        ("Peito Desossado",              desemp.get("peito_kg"),        "kg",    "—",     True),
        ("Rendimento Carcaça",           desemp.get("rend_carcaca_pct"),"%" ,   ">74%",  (desemp.get("rend_carcaca_pct") or 0) >= 74),
        ("Rendimento Peito",             desemp.get("rend_peito_pct"), "%" ,   ">30%",  (desemp.get("rend_peito_pct") or 0) >= 30),
    ]
    for i, (ind, val, uni, meta, ok) in enumerate(dados_bio):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        _val(ws1, r, 1, ind, bg=bg, align="left")
        _val(ws1, r, 2, val, "#,##0.000", bg=bg)
        _val(ws1, r, 3, uni, bg=bg)
        _val(ws1, r, 4, meta, bg=bg)
        _val(ws1, r, 5, "✓ OK" if ok else "⚠ Atenção",
             bg=C_VERDE2 if ok else C_RED)
    row += len(dados_bio) + 2

    row = _sec(ws1, row, "EPEF — European Production Efficiency Factor", ncols=8)
    _hdr(ws1, row, ["EPEF", "Classe", "Viabilidade (%)", f"Idade Abate (d)", "GPD (g/d)", "Eficiência Bio"], sz=9)
    r = row + 1
    ep = epef_d
    ecls_bg = {"EXCELENTE": C_VERDE2, "BOM": C_AMBER, "REGULAR": C_ORANGE, "RUIM": C_RED}
    _val(ws1, r, 1, ep["EPEF"],               "#,##0.0",  bg=ecls_bg.get(ep["classe"], C_BRANCO))
    _val(ws1, r, 2, ep["classe"],              bg=ecls_bg.get(ep["classe"], C_BRANCO))
    _val(ws1, r, 3, ep["viabilidade_pct"],     "0.0",      bg=C_VERDE2 if ep["viabilidade_pct"] >= 95 else C_AMBER)
    _val(ws1, r, 4, int(idade_d),              bg=C_BRANCO)
    _val(ws1, r, 5, ep["ganho_diario_g"],      "#,##0.0",  bg=C_BRANCO)
    _val(ws1, r, 6, ep["eficiencia_bio"],      "#,##0.000", bg=C_BRANCO)
    row += 3

    row = _sec(ws1, row, "RESULTADO ECONÔMICO  ·  MOFC / IOAC", ncols=8, bg=C_VERDE)
    _hdr(ws1, row, ["Cenário", "Receita ($/ave)", "Custo Feed ($/ave)",
                    "MOFC ($/ave)", "IOAC ($/ave)", "MOFC Frota ($)", "IOAC Frota ($)"], sz=9, bg=C_VERDE)
    cen_data = [
        ("Frango Vivo",    "mofc_vivo",      "receita_vivo",      f"mofc_vivo"),
        ("Carcaça WOG",    "mofc_carcaca",   "receita_carcaca",   f"mofc_carcaca"),
        ("Peito Desossado","mofc_desossado","receita_desossado", f"mofc_desossado"),
    ]
    n_vend = frota["n_aves_vendidas"]
    for i, (nm, k_mofc, k_rec, k_fr) in enumerate(cen_data):
        r = row + 1 + i
        bg = C_VERDE2 if i % 2 == 0 else C_BRANCO
        mofc_v = resultado.get(k_mofc, 0)
        ioac_v = mofc_v - custo_ext
        rec_v  = resultado.get(k_rec, 0)
        _val(ws1, r, 1, nm,             bg=bg, align="left")
        _val(ws1, r, 2, rec_v,          "$#,##0.0000", bg=bg)
        _val(ws1, r, 3, resultado.get("custo_alimentacao_por_ave"), "$#,##0.0000", bg=bg)
        _val(ws1, r, 4, mofc_v,         "$#,##0.0000", bg=C_VERDE2 if mofc_v > 0 else C_RED)
        _val(ws1, r, 5, ioac_v,         "$#,##0.0000", bg=C_VERDE2 if ioac_v > 0 else C_RED)
        _val(ws1, r, 6, mofc_v * n_vend, "$#,##0",     bg=bg)
        _val(ws1, r, 7, ioac_v * n_vend, "$#,##0",     bg=bg)
    row += len(cen_data) + 2

    row = _sec(ws1, row, "CUSTO TOTAL DE PRODUÇÃO", ncols=8)
    _hdr(ws1, row, ["Item de Custo", "$/ave", "% do Total"], sz=9)
    itens = cprod["itens"]
    for i, (k, v) in enumerate(itens.items()):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        _val(ws1, r, 1, k, bg=bg, align="left")
        _val(ws1, r, 2, v, "$#,##0.0000", bg=bg)
        _val(ws1, r, 3, round(v / cprod["total"] * 100, 1), "0.0", bg=bg)
    rt = row + len(itens) + 1
    _val(ws1, rt, 1, "TOTAL", bold=True, bg=C_ACENTO, align="left")
    _val(ws1, rt, 2, cprod["total"], "$#,##0.0000", bold=True, bg=C_ACENTO)
    _val(ws1, rt, 3, 100.0, "0.0", bold=True, bg=C_ACENTO)

    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 14
    ws1.column_dimensions["G"].width = 14

    # ════════════════════════════════════════════════════════════════
    # ABA 2 · PROGRAMA NUTRICIONAL
    # ════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Programa Nutricional")
    ws2.sheet_view.showGridLines = False
    _titulo_aba(ws2, "PROGRAMA NUTRICIONAL  ·  Composição por Fase",
                "Formulação LP HiGHS · Custo mínimo com restrições nutricionais", ncols=8)

    row = 4
    row = _sec(ws2, row, "NÍVEIS NUTRICIONAIS PRINCIPAIS", ncols=8)
    _hdr(ws2, row, ["Fase", "ME (kcal/kg)", "Dig. Lisina (%)", "Custo Ração ($/kg)",
                    "CAA Ponderado", "Custo Alim/kg PV", "Status LP"])
    for i, (fase, nome) in enumerate(zip(FASES, F_NOME)):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        niv  = programa.get(fase, {})
        form = form_res.get(fase, {})
        _val(ws2, r, 1, nome, bg=bg, align="left")
        _val(ws2, r, 2, niv.get("ame_n"),     "#,##0",    bg=bg)
        _val(ws2, r, 3, niv.get("dig_lys"),   "0.000",    bg=bg)
        _val(ws2, r, 4, form.get("custo_por_kg"), "$#,##0.0000", bg=bg)
        _val(ws2, r, 5, desemp.get("caa"),    "0.000",    bg=bg)
        _val(ws2, r, 6, resultado.get("custo_por_kg_pv"), "$#,##0.0000", bg=bg)
        _val(ws2, r, 7, form.get("status", "?"), bg=C_VERDE2 if form.get("status") == "OK" else C_RED)
    row += len(FASES) + 2

    row = _sec(ws2, row, "COMPOSIÇÃO POR INGREDIENTE  (% inclusão)", ncols=8)
    from modules.formulation import get_ingredient_names, CUSTOS_POR_KG
    nomes_ing = get_ingredient_names()
    all_ings = set()
    for fase in FASES:
        all_ings.update(form_res.get(fase, {}).get("composicao", {}).keys())
    all_ings = sorted(all_ings)
    _hdr(ws2, row, ["Ingrediente (código)", "Custo ($/kg)"] + F_SHT + ["Total (% pond.)"])
    for i, code in enumerate(all_ings):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        nome = nomes_ing.get(code, str(code))
        _val(ws2, r, 1, f"{nome}  [{code}]", bg=bg, align="left")
        _val(ws2, r, 2, CUSTOS_POR_KG.get(code, 0), "$#,##0.0000", bg=bg)
        total_pond = 0.0
        from modules.optimizer import PROP_CONSUMO
        for j, fase in enumerate(FASES):
            v = form_res.get(fase, {}).get("composicao", {}).get(code, 0.0)
            total_pond += v * PROP_CONSUMO[fase]
            _val(ws2, r, j + 3, v, "#,##0.00", bg=bg if v > 0 else C_CINZA)
        _val(ws2, r, 7, round(total_pond, 3), "#,##0.00", bg=bg)
    row += len(all_ings) + 2

    row = _sec(ws2, row, "PERFIL NUTRICIONAL CALCULADO", ncols=8)
    NUT_LBL = {
        "ame_n":"ME (kcal/kg)", "dig_lys":"Dig. Lisina (%)", "dig_met":"Dig. Metionina (%)",
        "dig_cys":"Dig. Cistina (%)", "dig_thr":"Dig. Treonina (%)", "dig_trp":"Dig. Triptofano (%)",
        "dig_val":"Dig. Valina (%)", "ca_total":"Cálcio Total (%)", "p_npp":"P Não-Fítico (%)",
        "sodium":"Sódio (%)", "chloride":"Cloreto (%)", "potassium":"Potássio (%)",
        "cp":"Proteína Bruta (%)", "cf":"Extrato Etéreo (%)",
    }
    _hdr(ws2, row, ["Nutriente"] + F_SHT)
    for i, (nut, lbl) in enumerate(NUT_LBL.items()):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        _val(ws2, r, 1, lbl, bg=bg, align="left")
        for j, fase in enumerate(FASES):
            calc = form_res.get(fase, {}).get("nutrientes_calculados", {}).get(nut)
            if calc is not None:
                _val(ws2, r, j + 2, calc, "#,##0.0000", bg=bg)
            else:
                _val(ws2, r, j + 2, "—", bg=C_CINZA)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    for col in range(3, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 14

    # ════════════════════════════════════════════════════════════════
    # ABA 3 · ANÁLISE ECONÔMICA
    # ════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Análise Econômica")
    ws3.sheet_view.showGridLines = False
    _titulo_aba(ws3, "ANÁLISE ECONÔMICA  ·  Break-Even · MOFC · IOAC",
                "Feed-only vs Custo Total  ·  Por cenário de venda", ncols=8)

    from modules.optimizer import break_even_preco, break_even_custo_total

    row = 4
    row = _sec(ws3, row, "BREAK-EVEN POR CENÁRIO  ·  Preço mínimo de venda ($/kg)", ncols=8)
    _hdr(ws3, row, ["Cenário", "Preço Atual ($/kg)",
                    "BE Feed ($/kg)", "Margem Feed (%)",
                    "BE Total ($/kg)", "Margem Total (%)",
                    "Custo Total/ave", "IOAC/ave"])
    cen_cfg = [
        ("Frango Vivo",    "vivo",      "frango_vivo"),
        ("Carcaça WOG",    "carcaca",   "carcaca"),
        ("Peito Desossado","desossado", "peito_desossado"),
    ]
    for i, (nm, cen, key_p) in enumerate(cen_cfg):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        be_f  = break_even_preco(programa, precos, cen)
        be_t  = break_even_custo_total(resultado, custos_extras, cen, precos)
        pa    = precos.get(key_p, 0)
        mofc_v = resultado.get(f"mofc_{cen}", 0)
        ioac_v = mofc_v - custo_ext
        _val(ws3, r, 1, nm, bg=bg, align="left")
        _val(ws3, r, 2, pa, "$#,##0.0000", bg=bg)
        bef = be_f.get("break_even")
        bet = be_t.get("break_even_total")
        mf  = be_f.get("margem_atual_pct")
        mt  = be_t.get("margem_atual_pct")
        _val(ws3, r, 3, bef if bef and bef > 0 else "< $0", "$#,##0.0000" if isinstance(bef, float) and bef > 0 else None, bg=bg)
        _val(ws3, r, 4, mf, "0.00", bg=C_VERDE2 if (mf and mf > 5) else (C_AMBER if mf else C_CINZA))
        _val(ws3, r, 5, bet if bet and bet > 0 else "< $0", "$#,##0.0000" if isinstance(bet, float) and bet > 0 else None, bg=bg)
        _val(ws3, r, 6, mt, "0.00", bg=C_VERDE2 if (mt and mt > 5) else (C_AMBER if mt else C_CINZA))
        _val(ws3, r, 7, be_t.get("custo_total_ave"), "$#,##0.0000", bg=bg)
        _val(ws3, r, 8, ioac_v, "$#,##0.0000", bg=C_VERDE2 if ioac_v > 0 else C_RED)
    row += len(cen_cfg) + 2

    row = _sec(ws3, row, "CUSTO TOTAL DISCRIMINADO  ($/ave)", ncols=8)
    _hdr(ws3, row, ["Item", "$/ave", "% Total"])
    itens_c = cprod["itens"]
    tot_c = cprod["total"]
    for i, (k, v) in enumerate(itens_c.items()):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        _val(ws3, r, 1, k, bg=bg, align="left")
        _val(ws3, r, 2, v, "$#,##0.0000", bg=bg)
        _val(ws3, r, 3, round(v / tot_c * 100, 1) if tot_c else 0, "0.0", bg=bg)
    rt = row + len(itens_c) + 1
    _val(ws3, rt, 1, "TOTAL", bold=True, bg=C_ACENTO, align="left")
    _val(ws3, rt, 2, tot_c,   "$#,##0.0000", bold=True, bg=C_ACENTO)
    _val(ws3, rt, 3, 100.0,   "0.0",         bold=True, bg=C_ACENTO)
    row += len(itens_c) + 3

    row = _sec(ws3, row, "CUSTO POR GRUPO DE INGREDIENTES  ·  Grower", ncols=8)
    grupos = resultado.get("custo_grupos_grower", {})
    _hdr(ws3, row, ["Grupo", "$/kg ração", "% do Custo Ração"])
    ck_grower = resultado.get("custos_racao_por_kg", {}).get("grower", 1.0)
    for i, (g, v) in enumerate(grupos.items()):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        _val(ws3, r, 1, g, bg=bg, align="left")
        _val(ws3, r, 2, v, "$#,##0.00000", bg=bg)
        _val(ws3, r, 3, round(v / ck_grower * 100, 1) if ck_grower else 0, "0.0", bg=bg)

    for col in ["A","B","C","D","E","F","G","H"]:
        ws3.column_dimensions[col].width = 16
    ws3.column_dimensions["A"].width = 22

    # ════════════════════════════════════════════════════════════════
    # ABA 4 · AMINOÁCIDOS & BALANÇO ELETROLÍTICO
    # ════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("AA & Balanço Eletrolítico")
    ws4.sheet_view.showGridLines = False
    _titulo_aba(ws4, "AMINOÁCIDOS E BALANÇO ELETROLÍTICO",
                "Razões relativas à Lisina  ·  Ideal Ross 308  ·  EB mEq/kg", ncols=8)

    from modules.optimizer import IDEAL_PROTEIN
    AA_META = {"met_pct_lys": 36, "metcys_pct_lys": 75, "thr_pct_lys": 67,
               "trp_pct_lys": 18, "val_pct_lys": 77}
    AA_LBL  = {"met_pct_lys": "Met/Lys (%)", "metcys_pct_lys": "Met+Cys/Lys (%)",
               "thr_pct_lys": "Thr/Lys (%)", "trp_pct_lys": "Trp/Lys (%)",
               "val_pct_lys": "Val/Lys (%)"}
    ratios_all = resultado.get("ratios_aa", {})

    row = 4
    row = _sec(ws4, row, "RAZÕES DE AMINOÁCIDOS  ·  % relativa à Dig. Lisina", ncols=8)
    _hdr(ws4, row, ["Aminoácido", "Meta Ross308 (%)",
                    "Starter", "Grower", "Finisher 1", "Finisher 2",
                    "Pior / Melhor"])
    for i, (aa_k, aa_lbl) in enumerate(AA_LBL.items()):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        meta = AA_META[aa_k]
        _val(ws4, r, 1, aa_lbl, bg=bg, align="left")
        _val(ws4, r, 2, meta, "0.0", bg=bg, bold=True)
        vals = []
        for j, fase in enumerate(FASES):
            v = ratios_all.get(fase, {}).get(aa_k, 0)
            vals.append(v)
            ok_aa = v >= meta * 0.98
            _val(ws4, r, j + 3, v, "0.0",
                 bg=C_VERDE2 if ok_aa else C_RED)
        _val(ws4, r, 7, f"↓{min(vals):.1f}  ↑{max(vals):.1f}", bg=bg)
    row += len(AA_LBL) + 2

    row = _sec(ws4, row, "BALANÇO ELETROLÍTICO  ·  Na+K-Cl  (mEq/kg)  ·  Ideal: 200-280", ncols=8)
    _hdr(ws4, row, ["Fase", "EB (mEq/kg)", "Status", "Na (%)", "K (%)", "Cl (%)", "Obs."])
    eb_fases = resultado.get("eb_fases", {})
    for i, (fase, fname) in enumerate(zip(FASES, F_NOME)):
        r = row + 1 + i
        nuts = form_res.get(fase, {}).get("nutrientes_calculados", {})
        eb = eb_fases.get(fase, 0)
        if 200 <= eb <= 280:
            eb_bg, eb_txt = C_VERDE2, "✓ Ideal"
        elif 160 <= eb <= 320:
            eb_bg, eb_txt = C_AMBER, "⚠ Aceitável"
        else:
            eb_bg, eb_txt = C_RED, "✗ Fora da faixa"
        _val(ws4, r, 1, fname, bg=C_CLARO if i % 2 == 0 else C_BRANCO, align="left")
        _val(ws4, r, 2, eb, "#,##0.0", bg=eb_bg)
        _val(ws4, r, 3, eb_txt, bg=eb_bg)
        _val(ws4, r, 4, nuts.get("sodium", 0), "0.000", bg=C_CLARO if i % 2 == 0 else C_BRANCO)
        _val(ws4, r, 5, nuts.get("potassium", 0), "0.000", bg=C_CLARO if i % 2 == 0 else C_BRANCO)
        _val(ws4, r, 6, nuts.get("chloride", 0), "0.000", bg=C_CLARO if i % 2 == 0 else C_BRANCO)
        _val(ws4, r, 7, "EB = Na/22.99 + K/39.1 - Cl/35.45  (×1000)" if i == 0 else "",
             bg=C_BRANCO)

    ws4.column_dimensions["A"].width = 26
    for col in ["B","C","D","E","F","G"]:
        ws4.column_dimensions[col].width = 16

    # ════════════════════════════════════════════════════════════════
    # ABA 5 · FROTA
    # ════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Escala de Frota")
    ws5.sheet_view.showGridLines = False
    _titulo_aba(ws5, f"ESCALA DE FROTA  ·  {int(n_aves):,} aves  ·  {mortalidade_pct:.1f}% mortalidade",
                "Resultado econômico projetado para o plantel completo", ncols=5)

    row = 4
    row = _sec(ws5, row, "DADOS DE PRODUÇÃO", ncols=5)
    _hdr(ws5, row, ["Item", "Valor", "Unidade"])
    fr_rows = [
        ("Aves Alojadas",     int(frota["n_aves_alojadas"]),  "aves"),
        ("Aves Vendidas",     int(frota["n_aves_vendidas"]),  "aves"),
        ("Mortalidade",       int(frota["n_aves_alojadas"] - frota["n_aves_vendidas"]), "aves"),
        ("Mortalidade (%)",   mortalidade_pct,                "%"),
        ("Biomassa Total",    round(frota["peso_total_kg"]/1000, 2), "toneladas"),
    ]
    for i, (it, vl, un) in enumerate(fr_rows):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        _val(ws5, r, 1, it, bg=bg, align="left")
        _val(ws5, r, 2, vl, "#,##0.00", bg=bg)
        _val(ws5, r, 3, un, bg=bg)
    row += len(fr_rows) + 2

    row = _sec(ws5, row, "RESULTADO FINANCEIRO POR CENÁRIO", ncols=5, bg=C_VERDE)
    _hdr(ws5, row, ["Cenário", "Receita ($)", "MOFC ($)", "IOAC ($)", "IOAC/ave ($)"], bg=C_VERDE)
    cen_frota = [
        ("Frango Vivo",    "receita_vivo",       "mofc_vivo"),
        ("Carcaça WOG",   "receita_carcaca",    "mofc_carcaca"),
        ("Peito Desossado","receita_desossado",  "mofc_desossado"),
    ]
    n_vend = frota["n_aves_vendidas"]
    for i, (nm, k_rec, k_mf) in enumerate(cen_frota):
        r = row + 1 + i
        bg = C_VERDE2 if i % 2 == 0 else C_BRANCO
        mfc_t = frota.get(k_mf, 0)
        iac_t = mfc_t - custo_ext * n_vend
        _val(ws5, r, 1, nm,                       bg=bg, align="left")
        _val(ws5, r, 2, frota.get(k_rec, 0),      "$#,##0",   bg=bg)
        _val(ws5, r, 3, mfc_t,                    "$#,##0",   bg=C_VERDE2 if mfc_t > 0 else C_RED)
        _val(ws5, r, 4, iac_t,                    "$#,##0",   bg=C_VERDE2 if iac_t > 0 else C_RED)
        _val(ws5, r, 5, round(iac_t / n_vend, 4), "$#,##0.0000", bg=bg)
    row += len(cen_frota) + 2

    row = _sec(ws5, row, "CUSTO TOTAL DA FROTA", ncols=5)
    _hdr(ws5, row, ["Item de Custo", "$/ave", "Frota Total ($)"])
    for i, (k, v) in enumerate(itens_c.items()):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        _val(ws5, r, 1, k, bg=bg, align="left")
        _val(ws5, r, 2, v, "$#,##0.0000", bg=bg)
        _val(ws5, r, 3, v * n_aves, "$#,##0", bg=bg)
    rt = row + len(itens_c) + 1
    _val(ws5, rt, 1, "TOTAL", bold=True, bg=C_ACENTO, align="left")
    _val(ws5, rt, 2, tot_c, "$#,##0.0000", bold=True, bg=C_ACENTO)
    _val(ws5, rt, 3, tot_c * n_aves, "$#,##0", bold=True, bg=C_ACENTO)

    ws5.column_dimensions["A"].width = 26
    for col in ["B","C","D","E"]:
        ws5.column_dimensions[col].width = 18

    # ════════════════════════════════════════════════════════════════
    # ABA 6 · CURVA DE CRESCIMENTO
    # ════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Curva Crescimento")
    ws6.sheet_view.showGridLines = False
    me_pond  = float(resultado.get("me_ponderado",  3050))
    lys_pond = float(resultado.get("lys_ponderada", 1.10))
    _titulo_aba(ws6,
                f"CURVA DE CRESCIMENTO  ·  ME={me_pond:.0f}  Lys={lys_pond:.3f}",
                "Escalada ao RSM Genesis42d_G · Referência Ross 308", ncols=7)

    from modules.biological import curva_crescimento
    cc = curva_crescimento(me_pond, lys_pond, int(idade_d))

    row = 4
    _sec(ws6, row, f"Fator de Escala vs Ross 308: {cc['escala_vs_padrao']:.3f}×  "
                   f" ·  Idade de abate: {int(idade_d)} dias", ncols=7)
    row += 1
    _hdr(ws6, row, ["Dia", "Semana", "Peso Vivo (kg)", "GPD (g/d)",
                    "Consumo Acum. (kg)", "CAA Acumulado", "Obs."])
    for i, dia in enumerate(cc["dias"]):
        r = row + 1 + i
        bg = C_AMBER if dia == int(idade_d) else (C_CLARO if dia % 14 == 0 else C_BRANCO)
        _val(ws6, r, 1, dia, bg=bg)
        _val(ws6, r, 2, f"S{dia//7}" if dia % 7 == 0 else "", bg=bg)
        _val(ws6, r, 3, cc["peso_kg"][i], "#,##0.000", bg=bg)
        _val(ws6, r, 4, cc["ganho_diario_g"][i], "#,##0.0", bg=bg)
        _val(ws6, r, 5, cc["consumo_acum_kg"][i], "#,##0.000", bg=bg)
        _val(ws6, r, 6, cc["caa_acumulado"][i], "#,##0.000", bg=bg)
        _val(ws6, r, 7, "← ABATE" if dia == int(idade_d) else "", bg=bg,
             bold=(dia == int(idade_d)))

    ws6.column_dimensions["A"].width = 8
    ws6.column_dimensions["B"].width = 10
    for col in ["C","D","E","F"]:
        ws6.column_dimensions[col].width = 18
    ws6.column_dimensions["G"].width = 12

    # ════════════════════════════════════════════════════════════════
    # ABA 7 · SHADOW PRICES
    # ════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Shadow Prices")
    ws7.sheet_view.showGridLines = False
    _titulo_aba(ws7, "SHADOW PRICES  ·  Custo Marginal das Restrições Ativas",
                "LP HiGHS · Quanto o custo da ração aumenta ao elevar o requisito em 1 unidade", ncols=7)

    row = 4
    for fase in FASES:
        sp = form_res.get(fase, {}).get("shadow_prices", {})
        if not sp:
            continue
        row = _sec(ws7, row, f"SHADOW PRICES  ·  {fase.upper()}", ncols=7)
        _hdr(ws7, row, ["Nutriente", "Meta", "Tipo",
                        "Shadow Price ($/ton + 1 un)", "Shadow Price ($/kg + 1 un)",
                        "Status", "Impacto"])
        sorted_sp = sorted(sp.items(), key=lambda x: -abs(x[1]["shadow_ton"]))
        for i, (nut, sv) in enumerate(sorted_sp):
            r = row + 1 + i
            bg = C_CLARO if i % 2 == 0 else C_BRANCO
            shp = sv["shadow_ton"]
            _val(ws7, r, 1, NUT_LBL.get(nut, nut), bg=bg, align="left")
            _val(ws7, r, 2, sv["meta"], "0.0000", bg=bg)
            _val(ws7, r, 3, sv["tipo"].upper(), bg=bg)
            _val(ws7, r, 4, shp, "+#,##0.000;-#,##0.000;0",
                 bg=C_RED if shp > 0.1 else (C_AMBER if shp > 0.02 else C_VERDE2))
            _val(ws7, r, 5, sv["shadow_kg"], "+0.0000000;-0.0000000;0.0000000", bg=bg)
            _val(ws7, r, 6, "BINDING", bg=C_ORANGE)
            _val(ws7, r, 7, "Alto" if abs(shp) > 0.1 else ("Médio" if abs(shp) > 0.02 else "Baixo"),
                 bg=C_RED if abs(shp) > 0.1 else C_AMBER)
        row += len(sorted_sp) + 2

    ws7.column_dimensions["A"].width = 28
    ws7.column_dimensions["B"].width = 10
    ws7.column_dimensions["C"].width = 10
    ws7.column_dimensions["D"].width = 26
    ws7.column_dimensions["E"].width = 22
    ws7.column_dimensions["F"].width = 12
    ws7.column_dimensions["G"].width = 10

    # ════════════════════════════════════════════════════════════════
    # ABA 8 · PARÂMETROS
    # ════════════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("Parâmetros")
    ws8.sheet_view.showGridLines = False
    _titulo_aba(ws8, "PARÂMETROS DO MODELO  ·  Premissas e Entradas",
                "Preços, programa nutricional e custos de produção", ncols=4)

    row = 4
    row = _sec(ws8, row, "PREÇOS DE VENDA", ncols=4)
    _hdr(ws8, row, ["Produto", "$/kg"])
    preco_rows = [
        ("Frango Vivo",     precos.get("frango_vivo", 1.00)),
        ("Carcaça WOG",     precos.get("carcaca", 2.20)),
        ("Peito Desossado", precos.get("peito_desossado", 4.50)),
        ("Cortes Diversos", precos.get("cortes_misc", 1.10)),
    ]
    for i, (nm, vl) in enumerate(preco_rows):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        _val(ws8, r, 1, nm, bg=bg, align="left")
        _val(ws8, r, 2, vl, "$#,##0.0000", bg=bg)
    row += len(preco_rows) + 2

    row = _sec(ws8, row, "CUSTOS EXTRAS DE PRODUÇÃO  ($/ave)", ncols=4)
    _hdr(ws8, row, ["Item", "$/ave", "% do Total Extras"])
    total_extras = sum(custos_extras.values())
    for i, (k, v) in enumerate(custos_extras.items()):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        _val(ws8, r, 1, k.replace("_", " ").title(), bg=bg, align="left")
        _val(ws8, r, 2, v, "$#,##0.0000", bg=bg)
        _val(ws8, r, 3, round(v / total_extras * 100, 1) if total_extras else 0, "0.0", bg=bg)
    rt = row + len(custos_extras) + 1
    _val(ws8, rt, 1, "TOTAL EXTRAS", bold=True, bg=C_ACENTO, align="left")
    _val(ws8, rt, 2, total_extras, "$#,##0.0000", bold=True, bg=C_ACENTO)
    row += len(custos_extras) + 3

    row = _sec(ws8, row, "PROGRAMA NUTRICIONAL", ncols=4)
    _hdr(ws8, row, ["Fase", "ME (kcal/kg)", "Dig. Lisina (%)", "Custo Ração ($/kg)"])
    for i, (fase, fname) in enumerate(zip(FASES, F_NOME)):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        niv = programa.get(fase, {})
        _val(ws8, r, 1, fname, bg=bg, align="left")
        _val(ws8, r, 2, niv.get("ame_n"), "#,##0", bg=bg)
        _val(ws8, r, 3, niv.get("dig_lys"), "0.000", bg=bg)
        _val(ws8, r, 4, form_res.get(fase, {}).get("custo_por_kg"), "$#,##0.0000", bg=bg)
    row += len(FASES) + 2

    row = _sec(ws8, row, "PARÂMETROS DE FROTA", ncols=4)
    _hdr(ws8, row, ["Parâmetro", "Valor"])
    frot_params = [
        ("Aves alojadas",       n_aves),
        ("Mortalidade (%)",     mortalidade_pct),
        ("Idade de abate (d)",  int(idade_d)),
        ("ME ponderado",        float(resultado.get("me_ponderado", 0))),
        ("Lys ponderada",       float(resultado.get("lys_ponderada", 0))),
    ]
    for i, (k, v) in enumerate(frot_params):
        r = row + 1 + i
        bg = C_CLARO if i % 2 == 0 else C_BRANCO
        _val(ws8, r, 1, k, bg=bg, align="left")
        _val(ws8, r, 2, v, "#,##0.000", bg=bg)

    ws8.column_dimensions["A"].width = 28
    ws8.column_dimensions["B"].width = 16
    ws8.column_dimensions["C"].width = 16
    ws8.column_dimensions["D"].width = 16

    # ── Salva em memória ──────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
